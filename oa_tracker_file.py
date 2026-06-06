import os
import io
import json
import re
import xmlrpc.client
import time
import pandas as pd
from datetime import datetime, timedelta
from dotenv import load_dotenv
from openpyxl.utils import get_column_letter

# from google.oauth2 import service_account
# from googleapiclient.discovery import build
# from googleapiclient.http import MediaIoBaseUpload

# ---------------- LOAD ENV ----------------
load_dotenv()
url = os.getenv("ODOO_URL")
db = os.getenv("ODOO_DB")
username = os.getenv("ODOO_USERNAME")
password = os.getenv("ODOO_PASSWORD")
api_key = os.getenv("ODOO_API_KEY")


SHEET_ID = "1F7epdshmtSM8iPmSgYTY9l7Hwsz4X0uuvShVi87s75o"
DRIVE_FOLDER_ID = "1AbwlHjnLuw7u0OCcJ4NIq7T9UB6SofW0"

# ---------------- GOOGLE DRIVE (SERVICE ACCOUNT) ----------------
GOOGLE_SERVICE_ACCOUNT_FILE = "Credentials.json"
DRIVE_SCOPES = ["https://www.googleapis.com/auth/drive"]

# ---------------- DYNAMIC DATE FILTER ----------------
FROM_DATE = "2026-01-01 00:00:00"
# TO_DATE = "2025-09-30 23:59:59"
# TO_DATE = "2025-12-31 23:59:59"
TO_DATE = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

# Small tolerance to handle rounding differences between OA subtotal and delivered value.
DELIVERED_VALUE_TOLERANCE = 1.0

# These line items are not physically delivered via Delivery operations, so they should
# not be counted when determining OA delivery/production status.
# Add/remove keywords to match your product naming.
NON_DELIVERABLE_ITEM_KEYWORDS = [
    "MOULD",
    "DOCUMENTATION CHARGE",
    "OTHERS CHARGE",
]

# ---------------- ODOO CONNECTION ----------------
common = xmlrpc.client.ServerProxy(f"{url}/xmlrpc/2/common")
uid = common.authenticate(db, username, api_key or password, {})
models = xmlrpc.client.ServerProxy(f"{url}/xmlrpc/2/object")


def execute_kw_with_retry(
    model_proxy,
    db,
    uid,
    password,
    model_name,
    method,
    args,
    kwargs=None,
    retries=5,
    delay=2,
):
    """Execute an Odoo XML-RPC call with retry for transient gateway errors."""
    if kwargs is None:
        kwargs = {}

    for attempt in range(1, retries + 1):
        try:
            return model_proxy.execute_kw(
                db, uid, password, model_name, method, args, kwargs
            )
        except xmlrpc.client.ProtocolError as e:
            errcode = getattr(e, "errcode", None)
            is_gateway_error = errcode in (502, 503, 504)
            if is_gateway_error and attempt < retries:
                wait_s = delay * (2 ** (attempt - 1))
                print(
                    f"⚠️ XML-RPC ProtocolError {errcode}: {e}. Retrying in {wait_s}s... ({attempt}/{retries})"
                )
                time.sleep(wait_s)
                continue
            raise


def get_existing_fields(model_name, candidate_fields):
    """Return subset of candidate_fields that exist on the given model."""
    try:
        field_info = models.execute_kw(
            db,
            uid,
            api_key or password,
            model_name,
            "fields_get",
            [candidate_fields],
            {"attributes": ["type"]},
        )
        return [field for field in candidate_fields if field in field_info]
    except Exception:
        return []


def get_field_relation(model_name, field_name):
    """If field is many2one, return its relation model name."""
    try:
        info = models.execute_kw(
            db,
            uid,
            api_key or password,
            model_name,
            "fields_get",
            [[field_name]],
            {"attributes": ["type", "relation"]},
        )
        field_info = info.get(field_name, {})
        if field_info.get("type") == "many2one":
            return field_info.get("relation")
    except Exception:
        return None
    return None


def discover_product_template_size_fields():
    """Discover product.template size fields for Inch/CM/MM.

    Tries known technical names first, then falls back to scanning fields_get labels.
    Returns a dict with keys: 'cm', 'inch', 'mm' mapping to field names or None.
    """
    field_map = {"cm": None, "inch": None, "mm": None}

    known = {"cm": "size_cm", "inch": "size_inch", "mm": "size_mm"}
    existing = set(get_existing_fields("product.template", list(known.values())))
    for key, fname in known.items():
        if fname in existing:
            field_map[key] = fname

    if all(field_map.values()):
        return field_map

    try:
        all_fields = models.execute_kw(
            db,
            uid,
            api_key or password,
            "product.template",
            "fields_get",
            [],
            {"attributes": ["string", "type"]},
        )
    except Exception:
        return field_map

    def _token_match(label: str, token: str) -> bool:
        return re.search(rf"\b{re.escape(token)}\b", label) is not None

    candidates = {"cm": [], "inch": [], "mm": []}
    for fname, meta in (all_fields or {}).items():
        label = str((meta or {}).get("string") or "").upper()
        fname_u = str(fname).upper()
        if "SIZE" not in label and "SIZE" not in fname_u:
            continue
        if "INCH" in fname_u or "INCH" in label or "IN" in fname_u and "INCH" in label:
            candidates["inch"].append(fname)
        if _token_match(label, "CM") or "_CM" in fname_u or fname_u.endswith("CM"):
            candidates["cm"].append(fname)
        if _token_match(label, "MM") or "_MM" in fname_u or fname_u.endswith("MM"):
            candidates["mm"].append(fname)

    def _pick(unit_key: str):
        if field_map[unit_key]:
            return
        unit_candidates = candidates.get(unit_key) or []
        if not unit_candidates:
            return
        unit_upper = unit_key.upper()
        preferred = [
            f
            for f in unit_candidates
            if "SIZE" in f.upper() and unit_upper in f.upper()
        ]
        field_map[unit_key] = (preferred or unit_candidates)[0]

    _pick("inch")
    _pick("cm")
    _pick("mm")
    return field_map


# ---------------- FETCH SALE ORDERS ----------------
order_domain = [
    ("sales_type", "=", "oa"),
    ("state", "=", "sale"),
    ("is_hold", "=", False),
    ("pi_type", "=", "regular"),
    ("date_order", ">=", FROM_DATE),
    ("date_order", "<=", TO_DATE),
    ("company_id", "in", [1, 3]),
]

order_fields_base = [
    "id",
    "name",
    "date_order",
    "partner_id",
    "buyer_name",
    "company_id",
    "user_id",
    "team_id",
    "order_line",
]

order_optional_fields = get_existing_fields(
    "sale.order", ["brand", "marketing_person", "marketing_team"]
)
order_fields = order_fields_base + order_optional_fields

orders = models.execute_kw(
    db,
    uid,
    api_key or password,
    "sale.order",
    "search_read",
    [order_domain],
    {"fields": order_fields},
)
print(f"✅ {len(orders)} sale orders fetched")
if not orders:
    exit("⚠️ No sale orders found")


# ---------------- HELPER FUNCTION ----------------
def safe_name(value):
    """Return name from [id,name] or str, else blank for False/None"""
    if isinstance(value, list) and len(value) > 1:
        return value[1]
    elif value:
        return str(value)
    return ""


# ---------------- FETCH BUYER BRAND GROUP (buyer_name/brand) ----------------
buyer_brand_map = {}
buyer_model = get_field_relation("sale.order", "buyer_name")

buyer_ids = []
for order in orders:
    buyer_value = order.get("buyer_name")
    if buyer_value and isinstance(buyer_value, list) and len(buyer_value) > 0:
        buyer_ids.append(buyer_value[0])
buyer_ids = list(set(buyer_ids))

if buyer_model and buyer_ids:
    buyer_brand_field = get_existing_fields(buyer_model, ["brand"])
    if buyer_brand_field:
        try:
            print(f"🔄 Fetching buyer brand group from {buyer_model}...")
            buyers = models.execute_kw(
                db,
                uid,
                api_key or password,
                buyer_model,
                "read",
                [buyer_ids],
                {"fields": ["id", "brand"]},
            )
            for buyer in buyers:
                buyer_id = buyer.get("id")
                if not buyer_id:
                    continue
                buyer_brand_map[buyer_id] = safe_name(buyer.get("brand"))
            print(f"✅ Buyer brand group fetched for {len(buyer_brand_map)} buyers")
        except Exception as e:
            print(f"⚠️ Warning: Could not fetch buyer brand group: {e}")
    else:
        print(f"ℹ️ Field 'brand' not found on buyer model {buyer_model}")
else:
    if not buyer_model:
        print("ℹ️ buyer_name relation not found (buyer brand group skipped)")
    elif not buyer_ids:
        print("ℹ️ No buyers found on orders (buyer brand group skipped)")


def _format_number(value):
    if value is None:
        return ""
    try:
        value_float = float(value)
    except Exception:
        return str(value)
    if value_float.is_integer():
        return str(int(value_float))
    return str(round(value_float, 3))


def _format_3dp(value):
    if value is None or value == "":
        return ""
    try:
        value_float = float(value)
    except Exception:
        return str(value)
    return f"{value_float:.3f}"


def _is_non_deliverable_item(item_name: str) -> bool:
    if not item_name:
        return False
    upper = str(item_name).upper()
    return any(keyword in upper for keyword in NON_DELIVERABLE_ITEM_KEYWORDS)


def _normalize_size_value(value):
    if value is None or value == "":
        return None
    try:
        val = float(value)
    except Exception:
        return None
    val = round(val, 3)
    if float(val).is_integer():
        return int(val)
    return val


def _extract_sizes_from_text(text: str, company_name: str):
    """Extract size numbers from a text, normalized per-company.

    Rules:
    - Company "Zipper": sizes may be in CM or Inch; normalize all to CM.
    - Company "Metal Trims": sizes are in MM; keep in MM.
    """
    if not text or not company_name:
        return []

    company_upper = str(company_name).strip().upper()
    text_upper = str(text).strip().upper()
    if not text_upper:
        return []

    sizes = []

    if company_upper == "ZIPPER":
        # Match explicit units like "15 CM", "6 INCH", "6 IN".
        for m in re.finditer(r"(\d+(?:\.\d+)?)\s*(CM|INCHES?|IN)\b", text_upper):
            raw_val = m.group(1)
            unit = m.group(2)
            try:
                num = float(raw_val)
            except Exception:
                continue
            if unit.startswith("IN"):
                num = num * 2.54
            sizes.append(_normalize_size_value(num))

        # Match inch symbol like 6" or 6″.
        for m in re.finditer(r"(\d+(?:\.\d+)?)\s*(?:\"|″|”)", text_upper):
            raw_val = m.group(1)
            try:
                num = float(raw_val) * 2.54
            except Exception:
                continue
            sizes.append(_normalize_size_value(num))

    elif company_upper == "METAL TRIMS":
        for m in re.finditer(r"(\d+(?:\.\d+)?)\s*MM\b", text_upper):
            raw_val = m.group(1)
            try:
                num = float(raw_val)
            except Exception:
                continue
            sizes.append(_normalize_size_value(num))

    # Filter Nones, dedupe, sort
    sizes = [s for s in sizes if s is not None]
    sizes = sorted(set(sizes), key=lambda x: float(x))
    return sizes


def _parse_local_date(date_value):
    """Parse Odoo datetime/date string and return local date.

    Notes:
    - Input is usually UTC-like string: YYYY-MM-DD HH:MM:SS
    - We apply a fixed offset (kept consistent with previous script behavior).
    """
    if not date_value:
        return None

    if isinstance(date_value, datetime):
        dt_value = date_value
    else:
        text = str(date_value).strip()
        if not text:
            return None

        dt_value = None
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
            try:
                dt_value = datetime.strptime(text, fmt)
                break
            except ValueError:
                continue
        if dt_value is None:
            return None

    adjusted = dt_value + timedelta(hours=6, minutes=0)
    return adjusted.date()


def upload_excel_bytes_to_drive(excel_bytes, filename, folder_id):
    try:
        from google.oauth2 import service_account
        from googleapiclient.discovery import build
        from googleapiclient.http import MediaIoBaseUpload
    except ImportError as e:
        raise RuntimeError(
            "Google Drive upload dependencies are not installed. Install 'google-api-python-client' and 'google-auth' to enable upload."
        ) from e

    creds = service_account.Credentials.from_service_account_file(
        GOOGLE_SERVICE_ACCOUNT_FILE,
        scopes=DRIVE_SCOPES,
    )
    drive_service = build("drive", "v3", credentials=creds)

    file_metadata = {"name": filename}
    if folder_id:
        file_metadata["parents"] = [folder_id]

    media = MediaIoBaseUpload(
        io.BytesIO(excel_bytes),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        resumable=True,
    )
    created = (
        drive_service.files()
        .create(
            body=file_metadata,
            media_body=media,
            fields="id,webViewLink",
            supportsAllDrives=True,
        )
        .execute()
    )
    return created


# ---------------- BATCH FETCH ORDER LINES ----------------
all_line_ids = [
    line_id for order in orders for line_id in (order.get("order_line") or [])
]

line_optional_fields = get_existing_fields(
    "sale.order.line", ["qty_invoiced", "sizein", "sizecm", "sizemm"]
)


def fetch_order_lines_batched(line_ids, batch_size=1000):
    """Fetch order lines in batches to avoid 502 errors"""
    all_lines = []

    # Process line_ids in batches
    for i in range(0, len(line_ids), batch_size):
        batch_line_ids = line_ids[i : i + batch_size]

        try:
            batch_lines = models.execute_kw(
                db,
                uid,
                api_key or password,
                "sale.order.line",
                "read",
                [batch_line_ids],
                {
                    "fields": [
                        "id",
                        "order_id",
                        "product_template_id",
                        "product_uom_qty",
                        "price_unit",
                        "price_subtotal",
                    ]
                    + line_optional_fields,
                },
            )
            all_lines.extend(batch_lines)
            print(
                f"  📝 Order lines batch {i//batch_size + 1}: {len(batch_lines)} records"
            )
        except Exception as e:
            print(f"  ⚠️ Order lines batch {i//batch_size + 1} failed: {e}")

    return all_lines


print("🔄 Fetching order lines in batches...")
lines = fetch_order_lines_batched(all_line_ids)
print(f"✅ {len(lines)} order lines fetched")

line_map = {}
for line in lines:
    line_map[line["id"]] = line


# ---------------- FETCH PRODUCT TEMPLATE FG CATEGORY ----------------
product_template_ids = set()
for line in lines:
    pt_val = line.get("product_template_id")
    if isinstance(pt_val, list) and pt_val:
        product_template_ids.add(pt_val[0])
    elif isinstance(pt_val, int):
        product_template_ids.add(pt_val)

product_template_fg_field = get_existing_fields("product.template", ["fg_categ_type"])
product_template_size_field_map = discover_product_template_size_fields()
product_template_size_fields = [
    f for f in product_template_size_field_map.values() if f
]
product_fg_map = {}
product_size_map = {}


def fetch_product_templates_batched(template_ids, fields, batch_size=1000):
    """Fetch product templates in batches."""
    if not template_ids or not fields:
        return []

    all_templates = []
    template_ids = list(template_ids)
    for i in range(0, len(template_ids), batch_size):
        batch_ids = template_ids[i : i + batch_size]
        try:
            batch_templates = models.execute_kw(
                db,
                uid,
                api_key or password,
                "product.template",
                "read",
                [batch_ids],
                {"fields": ["id"] + fields},
            )
            all_templates.extend(batch_templates)
            print(
                f"  🏷️ Product templates batch {i//batch_size + 1}: {len(batch_templates)} records"
            )
        except Exception as e:
            print(f"  ⚠️ Product templates batch {i//batch_size + 1} failed: {e}")
    return all_templates


template_fields_to_fetch = sorted(
    set((product_template_fg_field or []) + (product_template_size_fields or []))
)

if product_template_ids and template_fields_to_fetch:
    print("🔄 Fetching product.template fields...")
    templates = fetch_product_templates_batched(
        product_template_ids, template_fields_to_fetch
    )
    for template in templates:
        template_id = template.get("id")
        if not template_id:
            continue

        if "fg_categ_type" in template_fields_to_fetch:
            product_fg_map[template_id] = safe_name(template.get("fg_categ_type"))

        if product_template_size_fields:
            product_size_map[template_id] = {
                "cm": (
                    template.get(product_template_size_field_map.get("cm"))
                    if product_template_size_field_map.get("cm")
                    else None
                ),
                "inch": (
                    template.get(product_template_size_field_map.get("inch"))
                    if product_template_size_field_map.get("inch")
                    else None
                ),
                "mm": (
                    template.get(product_template_size_field_map.get("mm"))
                    if product_template_size_field_map.get("mm")
                    else None
                ),
            }

    if "fg_categ_type" in template_fields_to_fetch:
        print(f"✅ fg_categ_type fetched for {len(product_fg_map)} templates")
    if product_template_size_fields:
        print(f"✅ Size fields fetched for {len(product_size_map)} templates")
else:
    if not template_fields_to_fetch:
        print("ℹ️ No fg_categ_type/size fields found on product.template")
    else:
        print("ℹ️ No product templates found to read sizes")


# ---------------- FETCH PACKING RECORDS ----------------
# Get order IDs for packing/delivery lookup
order_ids = [order["id"] for order in orders]

packing_fields = [
    "id",
    "oa_id",
    "qty",
    "final_price",
    "action_date",
    "sale_order_line",
]


def fetch_packing_records_batched(order_ids, batch_size=500):
    """Fetch packing records in batches to avoid 502 errors"""
    all_packing_records = []

    base_packing_domain = [
        ("next_operation", "=", "FG Packing"),
        ("state", "!=", "done"),
        ("state", "!=", "closed"),
        ("company_id", "in", [1, 3]),
    ]

    # Process order_ids in batches
    for i in range(0, len(order_ids), batch_size):
        batch_order_ids = order_ids[i : i + batch_size]
        packing_domain_batch = base_packing_domain + [("oa_id", "in", batch_order_ids)]

        try:
            batch_records = models.execute_kw(
                db,
                uid,
                api_key or password,
                "operation.details",
                "search_read",
                [packing_domain_batch],
                {"fields": packing_fields},
            )
            all_packing_records.extend(batch_records)
            print(
                f"  📦 Packing batch {i//batch_size + 1}: {len(batch_records)} records"
            )
        except Exception as e:
            print(f"  ⚠️ Packing batch {i//batch_size + 1} failed: {e}")

    return all_packing_records


print("🔄 Fetching packing records in batches...")
packing_records = fetch_packing_records_batched(order_ids)
print(f"✅ {len(packing_records)} packing records fetched")

# ---------------- FETCH DELIVERY RECORDS ----------------
delivery_fields = packing_fields  # Same fields as packing


def fetch_delivery_records_batched(order_ids, batch_size=500):
    """Fetch delivery records in batches to avoid 502 errors"""
    all_delivery_records = []

    base_delivery_domain = [
        ("next_operation", "=", "Delivery"),
        ("state", "!=", "done"),
        ("state", "!=", "closed"),
        ("company_id", "in", [1, 3]),
    ]

    # Process order_ids in batches
    for i in range(0, len(order_ids), batch_size):
        batch_order_ids = order_ids[i : i + batch_size]
        delivery_domain = base_delivery_domain + [("oa_id", "in", batch_order_ids)]

        try:
            batch_records = models.execute_kw(
                db,
                uid,
                api_key or password,
                "operation.details",
                "search_read",
                [delivery_domain],
                {"fields": delivery_fields},
            )
            all_delivery_records.extend(batch_records)
            print(
                f"  🚚 Delivery batch {i//batch_size + 1}: {len(batch_records)} records"
            )
        except Exception as e:
            print(f"  ⚠️ Delivery batch {i//batch_size + 1} failed: {e}")

    return all_delivery_records


print("🔄 Fetching delivery records in batches...")
delivery_records = fetch_delivery_records_batched(order_ids)
print(f"✅ {len(delivery_records)} delivery records fetched")


def _extract_m2o_id(value):
    if isinstance(value, list) and value:
        return value[0]
    if isinstance(value, int):
        return value
    return None


def _to_float(value, default=0.0):
    try:
        if value is None or value == "":
            return default
        return float(value)
    except Exception:
        return default


def _date_qty_strings(date_qty_map):
    if not date_qty_map:
        return "", "", 0

    dates_sorted = sorted(date_qty_map.keys())
    qty_list = [date_qty_map[d] for d in dates_sorted]
    total_qty = sum(qty_list) if qty_list else 0
    return (
        ",".join(d.strftime("%d-%m-%Y") for d in dates_sorted),
        ",".join(_format_number(q) for q in qty_list),
        total_qty,
    )


def _date_qty_strings_with_sep(date_qty_map, inner_sep=","):
    if not date_qty_map:
        return "", "", 0

    dates_sorted = sorted(date_qty_map.keys())
    qty_list = [date_qty_map[d] for d in dates_sorted]
    total_qty = sum(qty_list) if qty_list else 0
    return (
        inner_sep.join(d.strftime("%d-%m-%Y") for d in dates_sorted),
        inner_sep.join(_format_number(q) for q in qty_list),
        total_qty,
    )


def fetch_invoiced_qty_by_oa(order_ids, batch_size=100):
    """Return {sale_order_id: invoiced_qty_sum} using combine.invoice.line.

    Returns:
        dict on success (may be empty if no invoices), or None if the call fails and
        the caller should fall back.
    """
    if not order_ids:
        return {}

    required_fields = get_existing_fields(
        "combine.invoice.line", ["sale_order", "quantity"]
    )
    if "sale_order" not in required_fields or "quantity" not in required_fields:
        print(
            "ℹ️ combine.invoice.line is missing sale_order/quantity; lc_status will use qty_invoiced fallback."
        )
        return None

    has_invoice_date = "invoice_date" in get_existing_fields(
        "combine.invoice.line", ["invoice_date"]
    )
    has_parent_state = "parent_state" in get_existing_fields(
        "combine.invoice.line", ["parent_state"]
    )

    base_domain = []
    if has_invoice_date:
        base_domain.extend(
            [("invoice_date", ">=", FROM_DATE), ("invoice_date", "<=", TO_DATE)]
        )

    # Align with sale.order.line.qty_invoiced behavior (posted invoices only).
    if has_parent_state:
        base_domain.append(("parent_state", "=", "posted"))

    qty_by_oa = {}
    total_batches = (len(order_ids) + batch_size - 1) // batch_size

    for i in range(0, len(order_ids), batch_size):
        batch_order_ids = order_ids[i : i + batch_size]
        domain = base_domain + [("sale_order", "in", batch_order_ids)]

        try:
            grouped = execute_kw_with_retry(
                models,
                db,
                uid,
                api_key or password,
                "combine.invoice.line",
                "read_group",
                [domain],
                {
                    "fields": ["sale_order", "quantity:sum"],
                    "groupby": ["sale_order"],
                    "lazy": False,
                },
            )
            print(
                f"  🧾 Invoice qty batch {i//batch_size + 1}/{total_batches}: {len(grouped or [])} groups"
            )
        except Exception as e:
            print(
                f"⚠️ Warning: Could not aggregate combine.invoice.line quantities (batch {i//batch_size + 1}/{total_batches}): {e}"
            )
            return None

        for row in grouped or []:
            oid = _extract_m2o_id(row.get("sale_order"))
            if not oid:
                continue
            qty_val = row.get("quantity")
            if qty_val is None:
                qty_val = row.get("quantity_sum")
            qty_by_oa[oid] = qty_by_oa.get(oid, 0.0) + _to_float(qty_val, 0.0)

    print(f"✅ Invoiced qty aggregated for {len(qty_by_oa)} OAs")
    return qty_by_oa


# ---------------- AGGREGATE PACKING/DELIVERY BY OA + ACTION DATE ----------------
# ---------------- FETCH OPERATION.DETAILS FG BALANCE (per OA) ----------------
op_fields_for_fg = ["id", "oa_id", "fg_balance", "final_price"]
op_fields_for_fg += get_existing_fields("operation.details", ["sale_order_line"])


def fetch_operation_details_batched(order_ids, batch_size=500):
    all_records = []
    # Use FG-specific domain (matches fg_stock.py): only records with FG balance
    base_domain = [
        ("next_operation", "=", "FG Packing"),
        ("state", "!=", "done"),
        ("state", "!=", "closed"),
        ("company_id", "in", [1, 3]),
        ("fg_balance", ">", 0),
    ]
    for i in range(0, len(order_ids), batch_size):
        batch_order_ids = order_ids[i : i + batch_size]
        domain = base_domain + [("oa_id", "in", batch_order_ids)]
        try:
            batch = models.execute_kw(
                db,
                uid,
                api_key or password,
                "operation.details",
                "search_read",
                [domain],
                {"fields": op_fields_for_fg},
            )
            all_records.extend(batch)
            print(
                f"  🔎 operation.details batch {i//batch_size + 1}: {len(batch)} records"
            )
        except Exception as e:
            print(f"  ⚠️ operation.details batch {i//batch_size + 1} failed: {e}")
    return all_records


print("🔄 Fetching operation.details FG balances in batches...")
op_records = fetch_operation_details_batched(order_ids)
print(f"✅ {len(op_records)} operation.details records fetched for FG balances")

fg_balance_by_oa = {}
fg_value_by_oa = {}
for r in op_records:
    oid = _extract_m2o_id(r.get("oa_id"))
    if not oid:
        continue
    fb = _to_float(r.get("fg_balance"), 0.0)
    fp = _to_float(r.get("final_price"), 0.0)
    fg_balance_by_oa[oid] = fg_balance_by_oa.get(oid, 0.0) + fb
    fg_value_by_oa[oid] = fg_value_by_oa.get(oid, 0.0) + (fb * fp)

# ---------------- FETCH MANUFACTURING ORDERS (PRODUCTION PENDING) ----------------
MANUF_MODEL = "manufacturing.order"
MANUF_FIELDS = [
    "id",
    "oa_id",
    "balance_qty",
    "fg_categ_type",
    "oa_total_balance",
    "sale_order_line",
    "state",
]


MANUF_DOMAIN = [
    ("fg_categ_type", "!=", ""),
    ("balance_qty", ">", 0),
    ("oa_total_balance", ">", 0),
    ("oa_id", "!=", False),
    ("state", "not in", ("closed", "cancel", "hold")),
    ("company_id", "in", [1, 3]),
    ("sale_order_line.state", "=", "sale"),
]


def fetch_manuf_orders():
    try:
        return models.execute_kw(
            db,
            uid,
            api_key or password,
            MANUF_MODEL,
            "search_read",
            [MANUF_DOMAIN],
            {"fields": MANUF_FIELDS},
        )
    except Exception as e:
        print(f"⚠️ manufacturing.order fetch failed: {e}")
        return []


print("🔄 Fetching manufacturing orders (production pending)...")
manuf_records = fetch_manuf_orders()
print(f"✅ {len(manuf_records)} manufacturing.order records fetched")

# ---------------- ITEM-WISE MAPPING (sale.order.line -> item name) ----------------
sale_line_ids = set()

for rec in packing_records:
    sol_id = _extract_m2o_id(rec.get("sale_order_line"))
    if sol_id:
        sale_line_ids.add(sol_id)

for rec in delivery_records:
    sol_id = _extract_m2o_id(rec.get("sale_order_line"))
    if sol_id:
        sale_line_ids.add(sol_id)

for rec in op_records:
    sol_id = _extract_m2o_id(rec.get("sale_order_line"))
    if sol_id:
        sale_line_ids.add(sol_id)

for rec in manuf_records:
    sol_id = _extract_m2o_id(rec.get("sale_order_line"))
    if sol_id:
        sale_line_ids.add(sol_id)

sale_line_ids = list(sale_line_ids)
sale_line_fields = get_existing_fields(
    "sale.order.line", ["product_template_id", "price_unit", "discount"]
)

sale_line_info = {}
if sale_line_ids and sale_line_fields:
    try:
        sale_lines = models.execute_kw(
            db,
            uid,
            api_key or password,
            "sale.order.line",
            "read",
            [sale_line_ids],
            {"fields": ["id"] + sale_line_fields},
        )
        for line in sale_lines:
            sol_id = line.get("id")
            if not sol_id:
                continue
            pt_val = line.get("product_template_id")
            pt_id = _extract_m2o_id(pt_val)
            pt_name = safe_name(pt_val)
            item_name = product_fg_map.get(pt_id, "") if pt_id else ""
            if not item_name:
                item_name = pt_name

            sale_line_info[sol_id] = {
                "item_name": item_name,
                "price_unit": _to_float(line.get("price_unit"), 0.0),
                "discount": _to_float(line.get("discount"), 0.0),
            }
    except Exception as e:
        print(f"⚠️ Failed to fetch sale.order.line details for item-wise mapping: {e}")


# ---------------- ITEM-WISE AGGREGATES ----------------
packing_by_oa_item = {}
for pack in packing_records:
    order_id = _extract_m2o_id(pack.get("oa_id"))
    if not order_id:
        continue
    action_date = _parse_local_date(pack.get("action_date"))
    if not action_date:
        continue
    sol_id = _extract_m2o_id(pack.get("sale_order_line"))
    item_name = sale_line_info.get(sol_id, {}).get("item_name")
    if not item_name:
        continue
    qty = _to_float(pack.get("qty"), 0.0)

    packing_by_oa_item.setdefault(order_id, {}).setdefault(item_name, {})
    packing_by_oa_item[order_id][item_name][action_date] = (
        packing_by_oa_item[order_id][item_name].get(action_date, 0.0) + qty
    )


delivery_qty_by_oa_item = {}
delivery_value_by_oa_item = {}
for delivery in delivery_records:
    order_id = _extract_m2o_id(delivery.get("oa_id"))
    if not order_id:
        continue
    action_date = _parse_local_date(delivery.get("action_date"))
    if not action_date:
        continue
    sol_id = _extract_m2o_id(delivery.get("sale_order_line"))
    item_name = sale_line_info.get(sol_id, {}).get("item_name")
    if not item_name:
        continue

    qty = _to_float(delivery.get("qty"), 0.0)
    final_price = _to_float(delivery.get("final_price"), 0.0)
    value = qty * final_price

    delivery_qty_by_oa_item.setdefault(order_id, {}).setdefault(item_name, {})
    delivery_qty_by_oa_item[order_id][item_name][action_date] = (
        delivery_qty_by_oa_item[order_id][item_name].get(action_date, 0.0) + qty
    )

    delivery_value_by_oa_item.setdefault(order_id, {}).setdefault(item_name, {})
    delivery_value_by_oa_item[order_id][item_name][action_date] = (
        delivery_value_by_oa_item[order_id][item_name].get(action_date, 0.0) + value
    )


fg_balance_by_oa_item = {}
fg_value_by_oa_item = {}
for r in op_records:
    oid = _extract_m2o_id(r.get("oa_id"))
    if not oid:
        continue
    sol_id = _extract_m2o_id(r.get("sale_order_line"))
    item_name = sale_line_info.get(sol_id, {}).get("item_name")
    if not item_name:
        continue

    fb = _to_float(r.get("fg_balance"), 0.0)
    fp = _to_float(r.get("final_price"), 0.0)
    fg_balance_by_oa_item.setdefault(oid, {})
    fg_value_by_oa_item.setdefault(oid, {})
    fg_balance_by_oa_item[oid][item_name] = (
        fg_balance_by_oa_item[oid].get(item_name, 0.0) + fb
    )
    fg_value_by_oa_item[oid][item_name] = fg_value_by_oa_item[oid].get(
        item_name, 0.0
    ) + (fb * fp)


prod_pending_qty_by_oa = {}
prod_pending_value_by_oa = {}
prod_pending_qty_by_oa_item = {}
prod_pending_value_by_oa_item = {}

for rec in manuf_records:
    oid = _extract_m2o_id(rec.get("oa_id"))
    if not oid:
        continue
    balance_qty_val = _to_float(rec.get("balance_qty"), 0.0)
    sol_val = rec.get("sale_order_line")
    sol_id = sol_val[0] if isinstance(sol_val, list) and sol_val else None
    sol_data = sale_line_info.get(sol_id, {})
    price_unit_val = _to_float(sol_data.get("price_unit"), 0.0)
    discount_val = _to_float(sol_data.get("discount"), 0.0)

    gross_pending_value = balance_qty_val * price_unit_val
    pending_value_val = gross_pending_value * (1 - (discount_val / 100))

    prod_pending_qty_by_oa[oid] = prod_pending_qty_by_oa.get(oid, 0.0) + balance_qty_val
    prod_pending_value_by_oa[oid] = (
        prod_pending_value_by_oa.get(oid, 0.0) + pending_value_val
    )

    item_name = sol_data.get("item_name")
    if item_name:
        prod_pending_qty_by_oa_item.setdefault(oid, {})
        prod_pending_value_by_oa_item.setdefault(oid, {})
        prod_pending_qty_by_oa_item[oid][item_name] = (
            prod_pending_qty_by_oa_item[oid].get(item_name, 0.0) + balance_qty_val
        )
        prod_pending_value_by_oa_item[oid][item_name] = (
            prod_pending_value_by_oa_item[oid].get(item_name, 0.0) + pending_value_val
        )


packing_by_oa = {}
for pack in packing_records:
    order_id = _extract_m2o_id(pack.get("oa_id"))
    if not order_id:
        continue
    action_date = _parse_local_date(pack.get("action_date"))
    if not action_date:
        continue
    qty = pack.get("qty") or 0
    packing_by_oa.setdefault(order_id, {})
    packing_by_oa[order_id][action_date] = (
        packing_by_oa[order_id].get(action_date, 0) + qty
    )

delivery_qty_by_oa = {}
delivery_value_by_oa = {}
for delivery in delivery_records:
    order_id = _extract_m2o_id(delivery.get("oa_id"))
    if not order_id:
        continue
    action_date = _parse_local_date(delivery.get("action_date"))
    if not action_date:
        continue
    qty = _to_float(delivery.get("qty"), 0.0)
    final_price = _to_float(delivery.get("final_price"), 0.0)
    value = qty * final_price

    delivery_qty_by_oa.setdefault(order_id, {})
    delivery_qty_by_oa[order_id][action_date] = (
        delivery_qty_by_oa[order_id].get(action_date, 0) + qty
    )

    delivery_value_by_oa.setdefault(order_id, {})
    delivery_value_by_oa[order_id][action_date] = (
        delivery_value_by_oa[order_id].get(action_date, 0) + value
    )


# ---------------- INVOICED QTY (FOR lc_status) ----------------
invoiced_qty_by_oa = fetch_invoiced_qty_by_oa(order_ids)
use_invoice_line_qty = invoiced_qty_by_oa is not None
if invoiced_qty_by_oa is None:
    invoiced_qty_by_oa = {}


# ---------------- BUILD FINAL OA-LEVEL DATA ----------------
all_data = []

for order in orders:
    order_id = order.get("id")
    oa = safe_name(order.get("name"))
    released_date = _parse_local_date(order.get("date_order"))
    customer_name = safe_name(order.get("partner_id"))
    buyer_name = safe_name(order.get("buyer_name"))
    buyer_id = None
    buyer_value = order.get("buyer_name")
    if buyer_value and isinstance(buyer_value, list) and len(buyer_value) > 0:
        buyer_id = buyer_value[0]
    brand = (
        buyer_brand_map.get(buyer_id, "") or safe_name(order.get("brand")) or buyer_name
    )
    company = safe_name(order.get("company_id"))
    salesperson = safe_name(order.get("user_id"))
    team = safe_name(order.get("team_id"))
    marketing_person = safe_name(order.get("marketing_person"))
    marketing_team = safe_name(order.get("marketing_team"))

    item_agg = {}
    size_by_item = {}
    total_invoiced_qty = (
        _to_float(invoiced_qty_by_oa.get(order_id), 0.0)
        if use_invoice_line_qty
        else 0.0
    )
    for line_id in order.get("order_line") or []:
        line = line_map.get(line_id)
        if not line:
            continue

        if not use_invoice_line_qty:
            total_invoiced_qty += _to_float(line.get("qty_invoiced"), 0.0)

        pt_id = _extract_m2o_id(line.get("product_template_id"))
        item_name = product_fg_map.get(pt_id, "") if pt_id else ""
        if not item_name:
            item_name = safe_name(line.get("product_template_id"))
        if not item_name:
            continue

        # Prefer explicit sale.order.line size fields (matches Odoo UI columns)
        company_upper = str(company).strip().upper()
        sizes_from_fields = []

        if "ZIPPER" in company_upper:
            cm_val = line.get("sizecm")
            inch_val = line.get("sizein")
            cm_num = _normalize_size_value(cm_val)
            if cm_num is not None:
                sizes_from_fields.append(cm_num)
            inch_num = _normalize_size_value(inch_val)
            if inch_num is not None:
                sizes_from_fields.append(_normalize_size_value(float(inch_num) * 2.54))

        elif "METAL TRIMS" in company_upper:
            mm_val = line.get("sizemm")
            mm_num = _normalize_size_value(mm_val)
            if mm_num is not None:
                sizes_from_fields.append(mm_num)

        for size_val in sizes_from_fields:
            if size_val is not None:
                size_by_item.setdefault(item_name, set()).add(size_val)

        # Fallback: parse from name text when size fields are missing/empty
        if not sizes_from_fields:
            product_name = safe_name(line.get("product_template_id"))
            for source_text in (product_name, item_name):
                for size_val in _extract_sizes_from_text(source_text, company):
                    size_by_item.setdefault(item_name, set()).add(size_val)

        qty = line.get("product_uom_qty") or 0
        value = line.get("price_subtotal") or 0

        if item_name not in item_agg:
            item_agg[item_name] = {
                "qty": 0.0,
                "value": 0.0,
                "unit_price_weighted_sum": 0.0,
                "unit_price_qty_sum": 0.0,
            }

        qty_float = _to_float(qty, 0.0)
        item_agg[item_name]["qty"] += qty_float
        item_agg[item_name]["value"] += _to_float(value, 0.0)

        price_unit = line.get("price_unit")
        if qty_float:
            item_agg[item_name]["unit_price_weighted_sum"] += (
                _to_float(price_unit, 0.0) * qty_float
            )
            item_agg[item_name]["unit_price_qty_sum"] += qty_float

    item_names = list(item_agg.keys())
    oa_qty_list = [item_agg[name]["qty"] for name in item_names]
    oa_value_list = [item_agg[name]["value"] for name in item_names]

    unit_price_list = []
    for name in item_names:
        qty_sum = item_agg[name].get("unit_price_qty_sum") or 0.0
        weighted = item_agg[name].get("unit_price_weighted_sum") or 0.0
        if qty_sum:
            unit_price_list.append(weighted / qty_sum)
        else:
            unit_price_list.append("")

    item_str = ",".join(item_names)
    oa_qty_str = ",".join(_format_number(x) for x in oa_qty_list)
    unit_price_str = ",".join(_format_3dp(x) for x in unit_price_list)
    oa_value_str = ",".join(_format_number(x) for x in oa_value_list)

    total_oa_qty = sum(oa_qty_list) if oa_qty_list else 0

    lc_status = ""
    if total_oa_qty:
        if total_invoiced_qty >= total_oa_qty:
            lc_status = "lcr"
        elif total_invoiced_qty == 0:
            lc_status = "lcp"
        else:
            lc_status = "lcrp"

    deliverable_item_names = [
        name for name in item_names if not _is_non_deliverable_item(name)
    ]
    deliverable_oa_qty_list = [item_agg[name]["qty"] for name in deliverable_item_names]
    deliverable_oa_value_list = [
        item_agg[name]["value"] for name in deliverable_item_names
    ]

    total_oa_qty_deliverable = (
        sum(deliverable_oa_qty_list) if deliverable_oa_qty_list else 0
    )
    total_oa_value_deliverable = (
        sum(_to_float(v, 0.0) for v in deliverable_oa_value_list)
        if deliverable_oa_value_list
        else 0.0
    )

    packed_dates, packed_qtys, total_packed_qty = _date_qty_strings(
        packing_by_oa.get(order_id, {})
    )
    delivery_dates, delivery_qtys, total_delivery_qty = _date_qty_strings(
        delivery_qty_by_oa.get(order_id, {})
    )
    _, delivery_values, total_delivery_value = _date_qty_strings(
        delivery_value_by_oa.get(order_id, {})
    )

    # ---------------- ITEM-WISE STRINGS (comma-aligned with item list) ----------------
    fg_balance_item_str = ",".join(
        _format_number(fg_balance_by_oa_item.get(order_id, {}).get(name))
        for name in item_names
    )
    fg_value_item_str = ",".join(
        _format_number(fg_value_by_oa_item.get(order_id, {}).get(name))
        for name in item_names
    )
    prod_pending_qty_item_str = ",".join(
        _format_number(prod_pending_qty_by_oa_item.get(order_id, {}).get(name))
        for name in item_names
    )
    prod_pending_value_item_str = ",".join(
        _format_number(prod_pending_value_by_oa_item.get(order_id, {}).get(name))
        for name in item_names
    )

    packed_qty_item_parts = []
    packed_date_item_parts = []
    for name in item_names:
        dates_s, qtys_s, _ = _date_qty_strings_with_sep(
            packing_by_oa_item.get(order_id, {}).get(name, {}),
            inner_sep=";",
        )
        packed_date_item_parts.append(dates_s)
        packed_qty_item_parts.append(qtys_s)
    packed_qty_item_str = ",".join(packed_qty_item_parts)
    packed_date_item_str = ",".join(packed_date_item_parts)

    delivery_qty_item_parts = []
    delivery_date_item_parts = []
    delivery_value_item_parts = []
    for name in item_names:
        d_dates_s, d_qtys_s, _ = _date_qty_strings_with_sep(
            delivery_qty_by_oa_item.get(order_id, {}).get(name, {}),
            inner_sep=";",
        )
        _, d_values_s, _ = _date_qty_strings_with_sep(
            delivery_value_by_oa_item.get(order_id, {}).get(name, {}),
            inner_sep=";",
        )
        delivery_date_item_parts.append(d_dates_s)
        delivery_qty_item_parts.append(d_qtys_s)
        delivery_value_item_parts.append(d_values_s)

    delivery_qty_item_str = ",".join(delivery_qty_item_parts)
    delivery_value_item_str = ",".join(delivery_value_item_parts)
    delivery_date_item_str = ",".join(delivery_date_item_parts)

    size_json = json.dumps(
        {
            name: sorted(size_by_item.get(name, set()), key=lambda x: float(x))
            for name in item_names
            if size_by_item.get(name)
        },
        ensure_ascii=False,
    )

    # ---------------- ITEM-WISE JSON (item -> value) ----------------
    oa_qty_json = json.dumps(
        {name: _format_number(item_agg[name].get("qty")) for name in item_names},
        ensure_ascii=False,
    )
    unit_price_json = json.dumps(
        {
            name: _format_3dp(unit_price_list[idx])
            for idx, name in enumerate(item_names)
        },
        ensure_ascii=False,
    )
    oa_value_json = json.dumps(
        {name: _format_number(item_agg[name].get("value")) for name in item_names},
        ensure_ascii=False,
    )

    fg_balance_json = json.dumps(
        {
            name: _format_number(fg_balance_by_oa_item.get(order_id, {}).get(name))
            for name in item_names
        },
        ensure_ascii=False,
    )
    fg_value_json = json.dumps(
        {
            name: _format_number(fg_value_by_oa_item.get(order_id, {}).get(name))
            for name in item_names
        },
        ensure_ascii=False,
    )

    prod_pending_qty_json = json.dumps(
        {
            name: _format_number(
                prod_pending_qty_by_oa_item.get(order_id, {}).get(name)
            )
            for name in item_names
        },
        ensure_ascii=False,
    )
    prod_pending_value_json = json.dumps(
        {
            name: _format_number(
                prod_pending_value_by_oa_item.get(order_id, {}).get(name)
            )
            for name in item_names
        },
        ensure_ascii=False,
    )

    packed_qty_json = json.dumps(
        {
            name: _date_qty_strings_with_sep(
                packing_by_oa_item.get(order_id, {}).get(name, {}), inner_sep=";"
            )[1]
            for name in item_names
        },
        ensure_ascii=False,
    )
    packed_date_json = json.dumps(
        {
            name: _date_qty_strings_with_sep(
                packing_by_oa_item.get(order_id, {}).get(name, {}), inner_sep=";"
            )[0]
            for name in item_names
        },
        ensure_ascii=False,
    )

    delivery_qty_json = json.dumps(
        {
            name: _date_qty_strings_with_sep(
                delivery_qty_by_oa_item.get(order_id, {}).get(name, {}), inner_sep=";"
            )[1]
            for name in item_names
        },
        ensure_ascii=False,
    )
    delivery_value_json = json.dumps(
        {
            name: _date_qty_strings_with_sep(
                delivery_value_by_oa_item.get(order_id, {}).get(name, {}),
                inner_sep=";",
            )[1]
            for name in item_names
        },
        ensure_ascii=False,
    )
    delivery_date_json = json.dumps(
        {
            name: _date_qty_strings_with_sep(
                delivery_qty_by_oa_item.get(order_id, {}).get(name, {}), inner_sep=";"
            )[0]
            for name in item_names
        },
        ensure_ascii=False,
    )

    total_prod_pending = _to_float(prod_pending_qty_by_oa.get(order_id, 0.0), 0.0)
    total_fg_balance = _to_float(fg_balance_by_oa.get(order_id, 0.0), 0.0)
    total_delivered = _to_float(total_delivery_qty, 0.0)

    status = ""

    # 1. Fully Delivered
    if total_oa_qty_deliverable > 0 and total_delivered >= total_oa_qty_deliverable:
        status = "Delivered"

    # 2. Partially Delivered
    elif total_delivered > 0:
        status = "Partially Delivered"

    # 3. In Production (must come BEFORE FG)
    elif total_prod_pending > 0 and total_prod_pending < total_oa_qty_deliverable:
        status = "In Production"

    # 4. FG Stock (only when production fully done)
    elif total_prod_pending <= 0 and total_fg_balance > 0:
        status = "FG Stock"

    # 5. Pending Production
    else:
        status = "Pending Production"

    all_data.append(
        {
            "oa_no": oa,
            "released_date": released_date,
            "customer_name": customer_name,
            "buyer_name": buyer_name,
            "brand": brand,
            "company": company,
            "salesperson": salesperson,
            "team": team,
            "marketing_person": marketing_person,
            "marketing_team": marketing_team,
            "item": item_str,
            "size": size_json,
            # Item-wise JSON maps (kept under the original column names)
            "oa_qty": oa_qty_json,
            "unit_price": unit_price_json,
            "fg_balance": fg_balance_json,
            "fg_value": fg_value_json,
            "prod_pending_qty": prod_pending_qty_json,
            "prod_pending_value": prod_pending_value_json,
            "oa_value": oa_value_json,
            "packed_qty": packed_qty_json,
            "packed_date": packed_date_json,
            "delivery_qty": delivery_qty_json,
            "delivery_value": delivery_value_json,
            "delivery_date": delivery_date_json,
            "status": status,
            "lc_status": lc_status,
        }
    )


# ---------------- CREATE DATAFRAME AND EXPORT ----------------
df = pd.DataFrame(all_data)

column_order = [
    "oa_no",
    "released_date",
    "customer_name",
    "buyer_name",
    "brand",
    "company",
    "salesperson",
    "team",
    "marketing_person",
    "marketing_team",
    "item",
    "size",
    "oa_qty",
    "unit_price",
    "fg_balance",
    "fg_value",
    "prod_pending_qty",
    "prod_pending_value",
    "oa_value",
    "packed_qty",
    "packed_date",
    "delivery_qty",
    "delivery_value",
    "delivery_date",
    "status",
    "lc_status",
]

if df.empty:
    print("⚠️ No data found")
else:
    df = df[column_order]
    df["released_date"] = pd.to_datetime(df["released_date"], errors="coerce")
    exported_time = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = f"OA_Tracker_{exported_time}.xlsx"

    excel_buffer = io.BytesIO()
    with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="OA_Tracker")

        worksheet = writer.sheets["OA_Tracker"]
        released_col_idx = df.columns.get_loc("released_date") + 1
        released_col_letter = get_column_letter(released_col_idx)
        for row in range(2, len(df) + 2):
            cell = worksheet[f"{released_col_letter}{row}"]
            if cell.value is not None and cell.value != "":
                cell.number_format = "DD-MM-YYYY"

    excel_bytes = excel_buffer.getvalue()

    # ---------------- GOOGLE DRIVE UPLOAD (DISABLED) ----------------
    # try:
    #     uploaded = upload_excel_bytes_to_drive(
    #         excel_bytes=excel_bytes,
    #         filename=output_file,
    #         folder_id=DRIVE_FOLDER_ID,
    #     )
    #     link_or_id = uploaded.get("webViewLink") or uploaded.get("id")
    #     print(f"☁️ Export complete! Uploaded to Google Drive: {link_or_id}")
    # except Exception as e:
    #     print(f"❌ Google Drive upload failed: {e}")
    #     print(
    #         "ℹ️ Make sure the Drive folder is shared with the service account email, and that google-api-python-client is installed."
    #     )

    # ---------------- LOCAL EXPORT OPTION (DISABLED) ----------------
    # If you want local export instead, uncomment this block.
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="OA_Tracker")

        worksheet = writer.sheets["OA_Tracker"]
        released_col_idx = df.columns.get_loc("released_date") + 1
        released_col_letter = get_column_letter(released_col_idx)
        for row in range(2, len(df) + 2):
            cell = worksheet[f"{released_col_letter}{row}"]
            if cell.value is not None and cell.value != "":
                cell.number_format = "DD-MM-YYYY"

    print(f"📂 Export complete! File saved as {output_file}")
    print(f"ℹ️ Total records: {len(df)}")
